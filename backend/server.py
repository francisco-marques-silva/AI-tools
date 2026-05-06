import os
import sys
import uuid
import json
import threading
import time
import io
import random
import subprocess
import tempfile
import shutil
import concurrent.futures
from pathlib import Path as _Path
from typing import Dict, Any, List, Optional

from fastapi import FastAPI, HTTPException, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from pydantic import BaseModel, Field
import requests

from backend.providers import call_llm, RateLimitError, env_key_for


load_dotenv()

app = FastAPI(title="Triage Backend", version="0.1.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class StartPayload(BaseModel):
    provider: str = "openai"
    model: str = Field(default="gpt-5")
    api_key: str | None = None
    study_synopsis: str
    inclusion_criteria: List[str] = []
    exclusion_criteria: List[str] = []
    filename: str = ""
    sheet: str = ""
    records: List[Dict[str, Any]]
    temperature: float | None = None
    params: Optional[Dict[str, Any]] = None


JOBS: Dict[str, Dict[str, Any]] = {}

# Rate limiting / concurrency / resilience knobs
# Defaults tuned for OpenAI Tier 3 (gpt-4o: 5 000 RPM; reasoning/gpt-5: 600 RPM).
# Override any value via environment variable without restarting the server.
MAX_RETRIES = int(os.getenv("OPENAI_MAX_RETRIES", "5"))
BASE_BACKOFF = float(os.getenv("OPENAI_BASE_BACKOFF", "1.0"))
CONCURRENT_WORKERS = int(os.getenv("CONCURRENT_WORKERS", "20"))  # Tier 3 starting concurrency
CONCURRENT_MAX     = int(os.getenv("CONCURRENT_MAX",     "40"))  # Tier 3 ceiling (AIMD grows up to here)
CONCURRENT_MIN     = int(os.getenv("CONCURRENT_MIN",     "2"))   # never drop below 2 even on 429s
RECORD_MAX_RETRIES = int(os.getenv("RECORD_MAX_RETRIES", "3"))   # per-record retry attempts on error
AIUP_AFTER         = int(os.getenv("AIUP_AFTER",         "5"))   # successes before +1 slot (faster ramp for Tier 3)


class AdaptiveSemaphore:
    """
    AIMD concurrency controller:
    - Additive Increase: after AIUP_AFTER consecutive successes, add 1 slot (up to max).
    - Multiplicative Decrease: on rate-limit, halve current slots (down to min).
    Threads block in acquire() when all slots are occupied.
    """
    def __init__(self, initial: int, min_workers: int = 1, max_workers: int = 15,
                 increase_after: int = 8):
        self._cond   = threading.Condition(threading.Lock())
        self.current = max(min_workers, min(initial, max_workers))
        self.min_w   = min_workers
        self.max_w   = max_workers
        self._active = 0
        self._streak = 0           # consecutive successes
        self._up_after = increase_after
        self.rate_limit_hits = 0   # lifetime 429 events
        self.total_successes = 0

    # ---- slot management ----

    def acquire(self):
        with self._cond:
            while self._active >= self.current:
                self._cond.wait(timeout=0.5)
            self._active += 1

    def release(self):
        with self._cond:
            self._active = max(0, self._active - 1)
            self._cond.notify_all()

    def __enter__(self):
        self.acquire()
        return self

    def __exit__(self, *args):
        self.release()

    # ---- feedback signals ----

    def on_success(self):
        with self._cond:
            self.total_successes += 1
            self._streak += 1
            if self._streak >= self._up_after and self.current < self.max_w:
                self.current += 1
                self._streak  = 0
                self._cond.notify_all()

    def on_rate_limit(self):
        """Halve current slots (multiplicative decrease)."""
        with self._cond:
            self.rate_limit_hits += 1
            self._streak  = 0
            self.current  = max(self.min_w, self.current // 2)
            self._cond.notify_all()

    # ---- introspection ----

    @property
    def stats(self) -> Dict[str, Any]:
        with self._cond:
            return {
                "current_concurrency": self.current,
                "active_calls": self._active,
                "rate_limit_hits": self.rate_limit_hits,
                "total_successes": self.total_successes,
            }


def build_prompt(synopsis: str, inc: List[str], exc: List[str], title: str, abstract: str) -> str:
    inc_lines = "\n".join(f"- {i}" for i in inc) if inc else "- (none provided)"
    exc_lines = "\n".join(f"- {e}" for e in exc) if exc else "- (none provided)"
    parts: List[str] = [
                "You are a knowledgeable AI assistant tasked with high-sensitivity title and abstract screening of a research article for a systematic review. Follow a step-by-step evaluation focusing on not missing any potentially relevant study.",
        "",
        f"Synopsis/PICO: {synopsis.strip()}",
        "",
        "Inclusion Criteria:",
        inc_lines,
        "",
        "Exclusion Criteria:",
        exc_lines,
        "",
        f"Study Title: {title or ''}",
        f"Study Abstract: {abstract or ''}",
        "",
        "Instructions:",
        "1. Identify PICO elements and type record from the title and abstract: determine the studied population (animals/population), intervention/exposure end type of record (review, systematic review, original research article, case report).",
        "2. Check each inclusion criterion against the information: for each inclusion criterion, assess whether the abstract suggests the study fulfills it. (Treat unspecified details as uncertain rather than negative.)",
        "3. Check each exclusion criterion: assess if any exclusion criterion is clearly met by the study.",
        "4. Perform the above reasoning internally – do not output these steps.",
        "",
        "Decision logic (high recall focus):",
        "- If all inclusion criteria are met met and no exclusion criteria apply, Include the study.",
        "- If any inclusion criterion is clearly unmet or any exclusion criterion is definitely met, decide Exclude",
        "- If there is any uncertainty (e.g. some PICO elements are unclear from the abstract) and no clear exclusion, mark as Maybe rather than risk wrongful exclusion.",
        "",
        "When in doubt, err on the side of inclusion (include or maybe).",
        "",
        "Output (JSON only): Return a single JSON object with keys:",
        "- decision: \"include\" | \"exclude\" | \"maybe\"",
        "- rationale: brief reason (<=12 words)",
        "- inclusion_evaluation: array of { \"criterion\": string, \"status\": \"met\"|\"unclear\"|\"unmet\" }",
        "- exclusion_evaluation: array of { \"criterion\": string, \"status\": \"met\"|\"unclear\"|\"unmet\" }",
        "No other text should be produced outside the JSON.",
        "",
        "Example format:",
        "{",
        "  \"decision\": \"maybe\",",
        "  \"rationale\": \"Population matches, but intervention details unclear from abstract\",",
        "  \"inclusion_evaluation\": [ { \"criterion\": \"population: adults with T2D\", \"status\": \"met\" } ],",
        "  \"exclusion_evaluation\": [ { \"criterion\": \"non-human study\", \"status\": \"unmet\" } ]",
        "}",
        "",
        "Now, based on the above criteria and the article's title/abstract, output the JSON decision.",
    ]
    return "\n".join(parts)


def worker(job_id: str):
    job = JOBS[job_id]
    provider = job.get("provider", "openai")
    api_key = job.get("api_key") or os.getenv(env_key_for(provider))
    if not api_key:
        job["status"] = "error"
        job["error"] = f"API key missing (provide api_key in request or set {env_key_for(provider)} env var)"
        return
    records = job["records"]
    total = len(records)
    job["total"] = total
    job["processed"] = 0
    results: List[Dict[str, Any]] = []
    model = job["model"]
    synopsis = job["study_synopsis"]
    inc = job["inclusion_criteria"]
    exc = job["exclusion_criteria"]
    params = job.get("params")

    # Per-record retry budget (can be overridden via params)
    record_max_retries = RECORD_MAX_RETRIES
    if params and params.get("record_max_retries"):
        try:
            record_max_retries = max(1, min(int(params["record_max_retries"]), 10))
        except (ValueError, TypeError):
            pass

    # Adaptive concurrency setup (AIMD)
    init_concurrency = CONCURRENT_WORKERS
    max_concurrency  = CONCURRENT_MAX
    min_concurrency  = CONCURRENT_MIN
    aiup_after       = AIUP_AFTER
    if params:
        if params.get("concurrency"):
            try:
                init_concurrency = max(1, int(params["concurrency"]))
            except (ValueError, TypeError):
                pass
        if params.get("concurrency_max"):
            try:
                max_concurrency = max(1, int(params["concurrency_max"]))
            except (ValueError, TypeError):
                pass
        if params.get("concurrent_min"):
            try:
                min_concurrency = max(1, int(params["concurrent_min"]))
            except (ValueError, TypeError):
                pass
        if params.get("aiup_after"):
            try:
                aiup_after = max(1, int(params["aiup_after"]))
            except (ValueError, TypeError):
                pass

    # API-level retry / backoff (overridable via params)
    api_max_retries = MAX_RETRIES
    api_base_backoff = BASE_BACKOFF
    if params:
        if params.get("max_retries"):
            try:
                api_max_retries = max(1, min(int(params["max_retries"]), 20))
            except (ValueError, TypeError):
                pass
        if params.get("base_backoff"):
            try:
                api_base_backoff = max(0.1, min(float(params["base_backoff"]), 10.0))
            except (ValueError, TypeError):
                pass

    adaptive = AdaptiveSemaphore(
        initial=init_concurrency,
        min_workers=min_concurrency,
        max_workers=max_concurrency,
        increase_after=aiup_after,
    )
    job["concurrency_stats"] = adaptive.stats

    lock = threading.Lock()

    def process_record(idx: int, rec: Dict[str, Any]):
        """Process a single record with per-record retry on failure."""
        if job.get("status") == "cancelled":
            return
        title = rec.get("title") or ""
        abstract = rec.get("abstract") or ""
        rid = rec.get("id", idx)
        prompt = build_prompt(synopsis, inc, exc, title, abstract)

        out: Optional[Dict[str, Any]] = None
        error_log: List[str] = []
        attempts = 0

        for attempt in range(1, record_max_retries + 1):
            if job.get("status") == "cancelled":
                return
            attempts = attempt
            try:
                with adaptive:
                    out = call_llm(provider, model, prompt, api_key, params=params,
                                   max_retries=api_max_retries, base_backoff=api_base_backoff)
                    # Validate: decision must be one of the expected values
                    decision = str(out.get("decision", "")).strip().lower()
                    if decision not in {"include", "exclude", "maybe"}:
                        raise ValueError(f"Unexpected decision value: {decision!r}")
                # Signal success to the concurrency controller
                adaptive.on_success()
                # Success – stop retrying
                break

            except RateLimitError as e:
                # Signal rate-limit to halve concurrency before retrying
                adaptive.on_rate_limit()
                err_msg = str(e)
                error_log.append(f"attempt {attempt} [rate-limit]: {err_msg[:120]}")
                out = None
                # Honour Retry-After from the API, fallback to exponential backoff
                wait = e.retry_after if e.retry_after > 0 else api_base_backoff * (2 ** (attempt - 1))
                time.sleep(min(wait + random.uniform(0, 0.5), 30.0))

            except Exception as e:
                err_msg = str(e)
                error_log.append(f"attempt {attempt}: {err_msg[:120]}")
                out = None
                if attempt < record_max_retries:
                    backoff = min(api_base_backoff * (2 ** (attempt - 1)) + random.uniform(0, 0.3), 15.0)
                    time.sleep(backoff)

            finally:
                # Refresh live stats after every attempt
                job["concurrency_stats"] = adaptive.stats

        if out is None:
            # All retries exhausted – use a safe fallback and flag the record
            out = {
                "decision": "maybe",
                "rationale": "screening failed after retries – manual review required",
                "inclusion_evaluation": [{"criterion": c, "status": "unclear"} for c in inc],
                "exclusion_evaluation": [{"criterion": c, "status": "unclear"} for c in exc],
            }

        # Ensure per-criterion arrays exist even if the model omitted them
        inc_eval = out.get("inclusion_evaluation") or []
        exc_eval = out.get("exclusion_evaluation") or []
        if not inc_eval and inc:
            inc_eval = [{"criterion": c, "status": "unclear"} for c in inc]
        if not exc_eval and exc:
            exc_eval = [{"criterion": c, "status": "unclear"} for c in exc]

        entry = {
            "id": rid,
            "title": title,
            "abstract": abstract,
            "screening_decision": out["decision"],
            "screening_reason": out["rationale"],
            "inclusion_evaluation": inc_eval,
            "exclusion_evaluation": exc_eval,
            "_retries": attempts - 1,          # 0 = succeeded on first try
            "_error_log": error_log or None,   # None when no errors occurred
        }
        with lock:
            results.append(entry)
            job["processed"] = len(results)
            job["results"] = list(results)

    # Launch all records into a thread pool (pool size = ceiling, semaphore is the real gate)
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_concurrency) as executor:
        futures = []
        for idx, rec in enumerate(records, start=1):
            if job.get("status") == "cancelled":
                break
            futures.append(executor.submit(process_record, idx, rec))
        for f in concurrent.futures.as_completed(futures):
            if job.get("status") == "cancelled":
                for pending in futures:
                    pending.cancel()
                break
            try:
                f.result()
            except Exception:
                pass

    job["concurrency_stats"] = adaptive.stats
    if job.get("status") != "cancelled":
        job["status"] = "done"


@app.post("/api/start")
def start_job(payload: StartPayload):
    if not payload.records:
        raise HTTPException(status_code=400, detail="empty records")
    job_id = str(uuid.uuid4())
    JOBS[job_id] = {
        "status": "running",
        "provider": payload.provider or "openai",
        "model": payload.model or "gpt-5",
        "api_key": payload.api_key,
        "study_synopsis": payload.study_synopsis or "",
        "inclusion_criteria": payload.inclusion_criteria or [],
        "exclusion_criteria": payload.exclusion_criteria or [],
        "filename": payload.filename,
        "sheet": payload.sheet,
        "records": payload.records,
        "params": payload.params or {},
        "processed": 0,
        "total": len(payload.records),
        "results": [],
    }
    th = threading.Thread(target=worker, args=(job_id,), daemon=True)
    th.start()
    return {"job_id": job_id}


@app.post("/api/cancel/{job_id}")
def cancel(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    job["status"] = "cancelled"
    return {"status": "cancelled"}

@app.get("/api/status/{job_id}")
def status(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job não encontrado")
    resp: Dict[str, Any] = {
        "status": job.get("status"),
        "processed": job.get("processed", 0),
        "total": job.get("total", 0),
        "filename": job.get("filename"),
    }
    if job.get("concurrency_stats"):
        resp["concurrency"] = job["concurrency_stats"]
    return resp


@app.get("/api/progress/{job_id}")
def progress(job_id: str):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job nÃ£o encontrado")

    def event_stream():
        last = -1
        while True:
            j = JOBS.get(job_id)
            if not j:
                yield f"data: {json.dumps({'status': 'error', 'detail': 'job missing'})}\n\n"
                break
            processed = j.get("processed", 0)
            total = j.get("total", 0)
            status = j.get("status")
            if processed != last or status != "running":
                payload = {"status": status, "processed": processed, "total": total}
                # attach concurrency stats when available
                if j.get("concurrency_stats"):
                    payload["concurrency"] = j["concurrency_stats"]
                # attach last result summary if present
                try:
                    results = j.get("results") or []
                    if results:
                        lr = results[-1]
                        payload["last"] = {
                            "id": lr.get("id"),
                            "decision": lr.get("screening_decision"),
                            "rationale": lr.get("screening_reason"),
                        }
                except Exception:
                    pass
                if status == "error" and j.get("error"):
                    payload["error"] = str(j.get("error"))
                yield f"data: {json.dumps(payload)}\n\n"
                last = processed
            if status in {"done", "error", "cancelled"}:
                break
            time.sleep(0.5)

    return StreamingResponse(event_stream(), media_type="text/event-stream")


@app.get("/api/partial/{job_id}")
def partial(job_id: str, since: int = 0, limit: int = 200):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    results = job.get("results") or []
    total = job.get("total", 0)
    processed = job.get("processed", 0)
    status = job.get("status")
    start = max(0, int(since))
    end = min(len(results), start + max(1, int(limit)))
    items = []
    for idx in range(start, end):
        r = results[idx]
        item: Dict[str, Any] = {
            "index": idx + 1,
            "id": r.get("id"),
            "decision": r.get("screening_decision"),
            "rationale": r.get("screening_reason"),
        }
        retries = r.get("_retries", 0)
        if retries:
            item["retries"] = retries
        items.append(item)
    payload = {
        "status": status,
        "processed": processed,
        "total": total,
        "items": items,
        "next": end,
    }
    if status == "error" and job.get("error"):
        payload["error"] = str(job.get("error"))
    return JSONResponse(payload)

@app.get("/api/errors/{job_id}")
def record_errors(job_id: str):
    """Return a summary of all records that required retries or ultimately failed."""
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    results = job.get("results") or []
    problematic = []
    for r in results:
        retries = r.get("_retries", 0)
        error_log = r.get("_error_log")
        if retries or error_log:
            problematic.append({
                "id": r.get("id"),
                "title": r.get("title", "")[:80],
                "screening_decision": r.get("screening_decision"),
                "screening_reason": r.get("screening_reason"),
                "retries": retries,
                "error_log": error_log,
            })
    failed = [p for p in problematic if p["error_log"] and len(p["error_log"]) == job.get("record_max_retries", RECORD_MAX_RETRIES)]
    return JSONResponse({
        "job_status": job.get("status"),
        "total_records": job.get("total", 0),
        "total_with_retries": len(problematic),
        "total_failed": len(failed),
        "records": problematic,
    })


@app.get("/api/result/{job_id}")
def result(job_id: str, format: str = "csv"):
    job = JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job nÃ£o encontrado")
    if job.get("status") != "done":
        raise HTTPException(status_code=400, detail="job nÃ£o finalizado")
    rows = job.get("results", [])
    if not rows:
        return JSONResponse({"rows": []})
    fieldnames = [
        "id",
        "title",
        "abstract",
        "screening_decision",
        "screening_reason",
        "inclusion_evaluation",
        "exclusion_evaluation",
        "_retries",
        "_error_log",
    ]

    if (format or "").lower() == "xlsx":
        try:
            from openpyxl import Workbook
        except Exception:
            raise HTTPException(status_code=500, detail="openpyxl not installed on server. Install with 'pip install openpyxl'.")
        wb = Workbook()
        ws = wb.active
        ws.title = "triage"
        ws.append(fieldnames)
        for r in rows:
            def _cell(val):
                if isinstance(val, (list, dict)):
                    try:
                        return json.dumps(val, ensure_ascii=False)
                    except Exception:
                        return str(val)
                return val
            ws.append([_cell(r.get(k, "")) for k in fieldnames])
        bio = io.BytesIO()
        wb.save(bio)
        data = bio.getvalue()
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=triage_{job_id}.xlsx"
            },
        )
    else:
        # CSV in memory
        import csv
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for r in rows:
            row_out = {}
            for k in fieldnames:
                v = r.get(k, "")
                if isinstance(v, (list, dict)):
                    try:
                        v = json.dumps(v, ensure_ascii=False)
                    except Exception:
                        v = str(v)
                row_out[k] = v
            writer.writerow(row_out)
        data = output.getvalue().encode("utf-8")
        return Response(
            content=data,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename=triage_{job_id}.csv"
            },
        )


@app.get("/api/health")
def health():
    return {"status": "ok"}


# ===========================================================================
#  REPORT GENERATION
# ===========================================================================

REPORT_JOBS: Dict[str, Dict[str, Any]] = {}

_REPORT_SCRIPT  = str(_Path(__file__).resolve().parent / "report" / "main.py")
_GRAPHIC_SCRIPT = str(_Path(__file__).resolve().parent / "report" / "graphic.py")


def report_worker(job_id: str):
    job = REPORT_JOBS[job_id]
    input_dir  = job["input_dir"]
    output_dir = job["output_dir"]
    charts_dir = str(_Path(output_dir) / "charts")
    job["charts_dir"] = charts_dir
    job["charts"] = []
    env = {**os.environ, "PYTHONIOENCODING": "utf-8", "PYTHONUTF8": "1"}
    try:
        # ── Step 1: run the report pipeline ──────────────────────────
        proc = subprocess.Popen(
            [sys.executable, _REPORT_SCRIPT,
             "--input_dir", input_dir, "--output_dir", output_dir],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace", env=env,
        )
        for line in proc.stdout:
            job["log"].append(line.rstrip("\n"))
        proc.wait()

        out_files = sorted(f.name for f in _Path(output_dir).glob("*") if f.is_file())
        job["output_files"] = out_files

        if proc.returncode != 0:
            job["status"] = "error"
            return

        # ── Step 2: generate charts from the XLSX ────────────────────
        xlsx_files = sorted(_Path(output_dir).glob("data_grafics_*.xlsx"))
        if xlsx_files:
            _Path(charts_dir).mkdir(exist_ok=True)
            job["log"].append("")
            job["log"].append("  Generating charts…")
            cp = subprocess.Popen(
                [sys.executable, _GRAPHIC_SCRIPT, str(xlsx_files[-1]), "-o", charts_dir],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, encoding="utf-8", errors="replace", env=env,
            )
            for line in cp.stdout:
                job["log"].append(line.rstrip("\n"))
            cp.wait()
            job["charts"] = sorted(f.name for f in _Path(charts_dir).glob("*.png"))

        job["status"] = "done"
    except Exception as exc:
        job["log"].append(f"Internal error: {exc}")
        job["output_files"] = []
        job["charts"] = []
        job["status"] = "error"


@app.post("/api/report/start")
async def report_start(files: list[UploadFile] = File(...)):
    job_id = str(uuid.uuid4())
    tmp = tempfile.mkdtemp(prefix="aireport_")
    input_dir = str(_Path(tmp) / "input")
    output_dir = str(_Path(tmp) / "output")
    _Path(input_dir).mkdir()
    _Path(output_dir).mkdir()

    saved: List[str] = []
    for f in files:
        safe_name = _Path(f.filename or "").name
        if not safe_name:
            continue
        dest = _Path(input_dir) / safe_name
        content = await f.read()
        dest.write_bytes(content)
        saved.append(safe_name)

    if not saved:
        shutil.rmtree(tmp, ignore_errors=True)
        raise HTTPException(status_code=400, detail="No files received.")

    REPORT_JOBS[job_id] = {
        "status": "running",
        "log": [],
        "input_dir": input_dir,
        "output_dir": output_dir,
        "tmp": tmp,
        "output_files": [],
        "uploaded_files": saved,
    }
    th = threading.Thread(target=report_worker, args=(job_id,), daemon=True)
    th.start()
    return {"job_id": job_id, "files": saved}


@app.get("/api/report/stream/{job_id}")
def report_stream(job_id: str):
    job = REPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="report job not found")

    def event_gen():
        sent = 0
        while True:
            log = job.get("log", [])
            while sent < len(log):
                yield f"data: {json.dumps({'type': 'log', 'line': log[sent]})}\n\n"
                sent += 1
            status = job.get("status", "running")
            if status in ("done", "error"):
                yield f"data: {json.dumps({'type': 'done', 'status': status, 'files': job.get('output_files', []), 'charts': job.get('charts', [])})}\n\n"
                break
            time.sleep(0.25)

    return StreamingResponse(event_gen(), media_type="text/event-stream")


@app.get("/api/report/chart/{job_id}/{filename}")
def report_chart_image(job_id: str, filename: str):
    job = REPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="job not found")
    charts_dir = job.get("charts_dir")
    if not charts_dir:
        raise HTTPException(status_code=404, detail="no charts available")
    safe_name = _Path(filename).name
    if not safe_name.endswith(".png"):
        raise HTTPException(status_code=400, detail="only PNG files")
    file_path = _Path(charts_dir) / safe_name
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="chart not found")
    return Response(content=file_path.read_bytes(), media_type="image/png")


@app.get("/api/report/tabledata/{job_id}")
def report_tabledata_sheets(job_id: str):
    job = REPORT_JOBS.get(job_id)
    if not job or job.get("status") != "done":
        raise HTTPException(status_code=404, detail="job not ready")
    xlsx_files = sorted(_Path(job["output_dir"]).glob("data_grafics_*.xlsx"))
    if not xlsx_files:
        return JSONResponse({"sheets": []})
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_files[-1]), read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
        return JSONResponse({"sheets": sheets, "xlsx": xlsx_files[-1].name})
    except Exception as e:
        return JSONResponse({"sheets": [], "error": str(e)})


@app.get("/api/report/tabledata/{job_id}/{sheet_name}")
def report_tabledata_sheet(job_id: str, sheet_name: str):
    job = REPORT_JOBS.get(job_id)
    if not job or job.get("status") != "done":
        raise HTTPException(status_code=404, detail="job not ready")
    xlsx_files = sorted(_Path(job["output_dir"]).glob("data_grafics_*.xlsx"))
    if not xlsx_files:
        raise HTTPException(status_code=404, detail="no XLSX found")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_files[-1]), read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=404, detail="sheet not found")
        ws = wb[sheet_name]
        rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if v is None else str(v)) for v in row])
            if len(rows) >= 502:
                break
        wb.close()
        if not rows:
            return JSONResponse({"headers": [], "rows": []})
        return JSONResponse({"headers": rows[0], "rows": rows[1:500]})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/report/download/{job_id}/{filename}")
def report_download(job_id: str, filename: str):
    job = REPORT_JOBS.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="report job not found")
    safe_name = _Path(filename).name
    file_path = _Path(job["output_dir"]) / safe_name
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail="file not found")
    data = file_path.read_bytes()
    if safe_name.endswith(".docx"):
        mt = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    elif safe_name.endswith(".xlsx"):
        mt = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        mt = "application/octet-stream"
    return Response(
        content=data,
        media_type=mt,
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


# Serve static frontend from current directory at root — must be last (catch-all mount)
import pathlib as _pathlib
app.mount("/", StaticFiles(directory=str(_pathlib.Path(__file__).parent.parent / "frontend"), html=True), name="static")
